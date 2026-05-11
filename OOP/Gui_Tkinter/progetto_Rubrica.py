import tkinter as tk
from openpyxl import Workbook, load_workbook
from tkinter import messagebox
import os

# Nome del file Excel
FILE_EXCEL = "rubrica_contatti.xlsx"

# Verifica se il file esiste, crea con intestazione
if not os.path.exists(FILE_EXCEL):
    wb = Workbook()
    ws = wb.active
    ws.title = "Contatti"
    ws.append(['Nome', 'Cognome', 'Data di Nascita', 'Numero di Telefono'])
    wb.save(FILE_EXCEL)

root = tk.Tk()

# titolo della form
root.title('Rubrica contatti')

# Dimenzione della form
root.geometry("400x350")

# lo sfondo della form
root.configure(background="#3B3303")

# Lalel + Entry per Nome
tk.Label(root, text='Nome: ', font=('Roboto', 12), bg='#3B3303', fg='#fff', anchor='w').pack(fill='x', padx=20, pady=10)
entry_nome = tk.Entry(root,  width=30, font=('Roboto', 14) )
entry_nome.pack(pady=5)


# Lalel + Entry per Cognome
tk.Label(root, text='Cognome: ', font=('Roboto', 12), bg='#3B3303', fg='#fff', anchor='w').pack(fill='x', padx=20)
entry_cognome = tk.Entry(root,  width=30, font=('Roboto', 14) )
entry_cognome.pack(pady=5)

# Lalel + Entry per la Data di Nascita
tk.Label(root, text='Data di Nascita (gg/mm/aaaa): ', font=('Roboto', 12), bg='#3B3303', fg='#fff', anchor='w').pack(fill='x', padx=20)
entry_data_nascita = tk.Entry(root,  width=30, font=('Roboto', 14) )
entry_data_nascita.pack(pady=5)

# Lalel + Entry per il Telefono
tk.Label(root, text='Numero di Telefono: ', font=('Roboto', 12), bg='#3B3303', fg='#fff', anchor='w').pack(fill='x', padx=20)
entry_telefono = tk.Entry(root,  width=30, font=('Roboto', 14) )
entry_telefono.pack(pady=5)

# ----- Funzione per salvare i dati
def salva_dati():
    nome = entry_nome.get()
    cognome = entry_cognome.get()
    data = entry_data_nascita.get()
    telefono = entry_telefono.get()


    wb = load_workbook(FILE_EXCEL)
    ws = wb.active
    ws.append([nome, cognome, data, telefono])
    wb.save(FILE_EXCEL)

    messagebox.showinfo(
            # titolo popup
            "Info Salvataggio",

            # messaggio popup
            "Il salvataggio è andato a buon fine! 🙂🙂"
        )
    
    svuotaCampi()

# ----- Funzione per svuotare i campi   
def svuotaCampi():
    entry_nome.delete(0, tk.END)
    entry_cognome.delete(0, tk.END)
    entry_data_nascita.delete(0, tk.END)
    entry_telefono.delete(0, tk.END)

# -------- BOTTONE SALVA -------------
button = tk.Button(
    root,
    text='Salva Contatto',
    font=('Roboto', 14),
    bg='white',
    command=salva_dati
)
button.pack(pady=20)

root.mainloop()

